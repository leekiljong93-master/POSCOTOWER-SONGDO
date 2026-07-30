# -*- coding: utf-8 -*-
"""
logic_calculator.py
─────────────────────────────────────────────────────────────
조달청 기준 원가계산서(갑지) / 세부내역서(을지) 산출 및
서식이 적용된(Styled) 견적서 엑셀 생성 모듈
"""

import io
import re
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════
# 0. 서식 상수 (Design Token) — 여기만 바꾸면 전체 톤 일괄 변경
# ═══════════════════════════════════════════════════════════
FONT_NAME = "맑은 고딕"

C_TITLE_BG  = "1F3864"   # 제목 배경 (진한 남색)
C_HEADER_BG = "2E5C8A"   # 표 헤더 배경 (남색)
C_GROUP_BG  = "EDF2F8"   # 대분류(1.재료비 등) 배경
C_SUB_BG    = "DCE6F1"   # 소계(▶) 배경
C_TOTAL_BG  = "FFE699"   # 총계(■) 배경
C_ZEBRA_BG  = "F7F9FC"   # 짝수행 음영
C_BORDER    = "8EA9C1"   # 테두리
C_NOTE      = "808080"   # 비고/각주 글자

NUM_FMT = '#,##0;[Red]-#,##0;"-"'
QTY_FMT = '#,##0.##;[Red]-#,##0.##;"-"'

THIN   = Side(style="thin", color=C_BORDER)
MEDIUM = Side(style="medium", color="1F3864")
BOX    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 금액/수량 서식을 적용할 컬럼 키워드
AMOUNT_KEYS = ("금액", "합계", "단가", "비용", "원가", "재료비", "노무비", "경비", "소계", "계")
QTY_KEYS    = ("수량", "규격수량", "인원", "공수", "일수")


# ═══════════════════════════════════════════════════════════
# 1. 원가계산 순수 연산 로직 (기존 로직 그대로 유지)
# ═══════════════════════════════════════════════════════════
def calculate_cost_summary(df_calc, rates):
    """조달청 기준 원가계산서(갑지) 산출 순수 연산 로직"""
    direct_material = df_calc[df_calc['구분'] == '자재']['합계'].sum() if not df_calc.empty else 0
    direct_labor    = df_calc[df_calc['구분'] == '노무']['합계'].sum() if not df_calc.empty else 0
    equipment_exp   = df_calc[df_calc['구분'] == '장비']['합계'].sum() if not df_calc.empty else 0
    direct_cost_total = direct_material + direct_labor + equipment_exp

    indirect_labor = int(direct_labor * (rates['indirect_labor'] / 100))
    total_labor    = direct_labor + indirect_labor

    sanjae_ins    = int(direct_labor * (rates['sanjae'] / 100))
    goyong_ins    = int(direct_labor * (rates['goyong'] / 100))
    health_ins    = int(direct_labor * (rates['health'] / 100))
    elderly_ins   = int(health_ins * (rates['elderly'] / 100))
    pension_ins   = int(direct_labor * (rates['pension'] / 100))
    retire_deduct = int(direct_labor * (rates['retire'] / 100))
    env_cost      = int(direct_cost_total * (rates['env'] / 100))
    safety_base   = int((direct_material + direct_labor) * (rates['safety'] / 100))

    safety_mgt = 0 if direct_cost_total < 20000000 else safety_base
    safety_calc_reason = "총 공사금액 2천만 원 미만 제외" if direct_cost_total < 20000000 \
        else f"(재료비+직접노무비) × {rates['safety']}%"

    etc_exp = int((direct_material + total_labor) * (rates['etc_exp'] / 100))
    total_expense = (equipment_exp + sanjae_ins + goyong_ins + health_ins + elderly_ins
                     + pension_ins + retire_deduct + safety_mgt + env_cost + etc_exp)

    net_construction_cost = direct_material + total_labor + total_expense
    general_admin = int(net_construction_cost * (rates['general_admin'] / 100))
    profit = int((total_labor + total_expense + general_admin) * (rates['profit'] / 100))

    supply_value = net_construction_cost + general_admin + profit
    vat = int(supply_value * (rates['tax'] / 100))
    total_contract_price = supply_value + vat

    return [
        {"비목": "1. 재료비", "금액(원)": f"{direct_material:,}", "산출근거": "직접재료비 합계"},
        {"비목": " └ 직접재료비", "금액(원)": f"{direct_material:,}", "산출근거": "직접재료비 총액"},
        {"비목": "2. 노무비", "금액(원)": f"{total_labor:,}", "산출근거": "직접노무비 + 간접노무비"},
        {"비목": " └ 직접노무비", "금액(원)": f"{direct_labor:,}", "산출근거": "직접노무비 총액"},
        {"비목": " └ 간접노무비", "금액(원)": f"{indirect_labor:,}", "산출근거": f"직접노무비 × {rates['indirect_labor']}%"},
        {"비목": "3. 경비", "금액(원)": f"{total_expense:,}", "산출근거": "기계경비 + 제보험료 + 제비용 + 기타경비"},
        {"비목": " └ 기계경비(장비비)", "금액(원)": f"{equipment_exp:,}", "산출근거": "기계경비 총액"},
        {"비목": " └ 산재보험료", "금액(원)": f"{sanjae_ins:,}", "산출근거": f"직접노무비 × {rates['sanjae']}%"},
        {"비목": " └ 고용보험료", "금액(원)": f"{goyong_ins:,}", "산출근거": f"직접노무비 × {rates['goyong']}%"},
        {"비목": " └ 국민건강보험료", "금액(원)": f"{health_ins:,}", "산출근거": f"직접노무비 × {rates['health']}%"},
        {"비목": " └ 노인장기요양보험료", "금액(원)": f"{elderly_ins:,}", "산출근거": f"국민건강보험료 × {rates['elderly']}%"},
        {"비목": " └ 국민연금보험료", "금액(원)": f"{pension_ins:,}", "산출근거": f"직접노무비 × {rates['pension']}%"},
        {"비목": " └ 퇴직공제부금비", "금액(원)": f"{retire_deduct:,}", "산출근거": f"직접노무비 × {rates['retire']}%"},
        {"비목": " └ 산업안전보건관리비", "금액(원)": f"{safety_mgt:,}", "산출근거": safety_calc_reason},
        {"비목": " └ 환경보전비", "금액(원)": f"{env_cost:,}", "산출근거": f"직접공사비 × {rates['env']}%"},
        {"비목": " └ 기타경비", "금액(원)": f"{etc_exp:,}", "산출근거": f"(재료비+노무비) × {rates['etc_exp']}%"},
        {"비목": "▶ 순공사원가 (1+2+3)", "금액(원)": f"{net_construction_cost:,}", "산출근거": "재료비 + 노무비 + 경비"},
        {"비목": "4. 일반관리비", "금액(원)": f"{general_admin:,}", "산출근거": f"순공사원가 × {rates['general_admin']}%"},
        {"비목": "5. 이윤", "금액(원)": f"{profit:,}", "산출근거": f"(노무비+경비+일반관리비) × {rates['profit']}%"},
        {"비목": "▶ 공급가액", "금액(원)": f"{supply_value:,}", "산출근거": "순공사원가 + 일반관리비 + 이윤"},
        {"비목": "6. 부가가치세", "금액(원)": f"{vat:,}", "산출근거": f"공급가액 × {rates['tax']}%"},
        {"비목": "■ 총 공사예정금액(도급액)", "금액(원)": f"{total_contract_price:,}", "산출근거": "공급가액 + 부가가치세"},
    ]


# ═══════════════════════════════════════════════════════════
# 2. 서식 공통 유틸
# ═══════════════════════════════════════════════════════════
def _to_number(v):
    """'1,234,000' → 1234000 변환 (숫자 아니면 원본 반환)"""
    if v is None or (isinstance(v, (int, float)) and not isinstance(v, bool)):
        return v
    s = str(v).strip()
    if s in ("", "-", "nan", "None"):
        return 0 if s == "-" else v
    if re.fullmatch(r"-?[\d,]+(\.\d+)?", s):
        try:
            n = float(s.replace(",", ""))
            return int(n) if n.is_integer() else n
        except ValueError:
            return v
    return v


def _disp_len(v):
    """한글(2칸) 고려한 표시 폭 계산"""
    s = "" if v is None else str(v)
    return sum(2 if ord(ch) > 0x2000 else 1 for ch in s)


def _autofit(ws, header_row, min_w=9, max_w=48, extra=4):
    """열 너비 자동 조정"""
    for col in range(1, ws.max_column + 1):
        width = 0
        for row in range(header_row, ws.max_row + 1):
            width = max(width, _disp_len(ws.cell(row=row, column=col).value))
        ws.column_dimensions[get_column_letter(col)].width = \
            max(min_w, min(max_w, width + extra))


def _title_block(ws, ncols, title, subtitle_pairs):
    """상단 제목 블록 생성 → 표 헤더가 시작될 행 번호 반환"""
    last = get_column_letter(max(ncols, 3))

    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]
    c.value = title
    c.font = Font(name=FONT_NAME, size=18, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=C_TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    row = 2
    for label, value in subtitle_pairs:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(name=FONT_NAME, size=10, bold=True)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=max(ncols, 3))
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = Font(name=FONT_NAME, size=10)
        vc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

    ws.row_dimensions[row].height = 8   # 여백 행
    return row + 1


def _write_header(ws, row, columns):
    for j, name in enumerate(columns, start=1):
        c = ws.cell(row=row, column=j, value=str(name))
        c.font = Font(name=FONT_NAME, size=10.5, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=C_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(left=THIN, right=THIN, top=MEDIUM, bottom=THIN)
    ws.row_dimensions[row].height = 26


def _fmt_for(col_name):
    n = str(col_name)
    if any(k in n for k in QTY_KEYS):
        return QTY_FMT
    if any(k in n for k in AMOUNT_KEYS):
        return NUM_FMT
    return None


def _page_setup(ws, header_row, landscape=False):
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.print_options.horizontalCentered = True
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


# ═══════════════════════════════════════════════════════════
# 3. 시트별 작성 로직
# ═══════════════════════════════════════════════════════════
def _build_summary_sheet(ws, summary_df, project_name, writer_name, note=None):
    """원가계산서(갑지)"""
    cols = list(summary_df.columns)
    today = datetime.now().strftime("%Y. %m. %d.")
    hr = _title_block(
        ws, len(cols), "원 가 계 산 서 (갑지)",
        [("공 사 명", project_name), ("작성일자", today), ("작 성 자", writer_name)]
    )
    _write_header(ws, hr, cols)

    r = hr + 1
    for _, rec in summary_df.iterrows():
        label = str(rec[cols[0]])
        is_total = label.strip().startswith("■")
        is_sub   = label.strip().startswith("▶")
        is_child = label.strip().startswith("└")
        is_group = bool(re.match(r"^\d+\.", label.strip()))

        for j, cname in enumerate(cols, start=1):
            val = label if j == 1 else _to_number(rec[cname])
            c = ws.cell(row=r, column=j, value=val)
            c.border = BOX
            fmt = _fmt_for(cname)
            if fmt and isinstance(val, (int, float)):
                c.number_format = fmt

            # 정렬
            if j == 1:
                c.alignment = Alignment(horizontal="left", vertical="center",
                                        indent=2 if is_child else 0)
            elif isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # 행 유형별 서식
            if is_total:
                c.font = Font(name=FONT_NAME, size=12, bold=True, color="1F3864")
                c.fill = PatternFill("solid", start_color=C_TOTAL_BG)
                c.border = Border(left=THIN, right=THIN, top=MEDIUM, bottom=MEDIUM)
            elif is_sub:
                c.font = Font(name=FONT_NAME, size=11, bold=True, color="1F3864")
                c.fill = PatternFill("solid", start_color=C_SUB_BG)
            elif is_group:
                c.font = Font(name=FONT_NAME, size=10.5, bold=True)
                c.fill = PatternFill("solid", start_color=C_GROUP_BG)
            else:
                c.font = Font(name=FONT_NAME, size=10, color="333333")

            # 산출근거(마지막 열)는 회색 소형 글씨
            if j == len(cols) and not (is_total or is_sub):
                c.font = Font(name=FONT_NAME, size=9, color=C_NOTE)

        ws.row_dimensions[r].height = 22 if (is_total or is_sub) else 18
        r += 1

    # 각주
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(len(cols), 3))
    nc = ws.cell(row=r, column=1, value=note or
                 "※ 본 산출내역은 조달청 「원가계산 제비율 적용기준」에 따라 작성되었으며, "
                 "적용요율 변동 시 금액이 변경될 수 있습니다.")
    nc.font = Font(name=FONT_NAME, size=9, italic=True, color=C_NOTE)
    nc.alignment = Alignment(horizontal="left", vertical="center")

    _autofit(ws, hr)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 30)
    _page_setup(ws, hr)


def _build_estimate_sheet(ws, estimate_df, project_name, writer_name):
    """세부내역서(을지)"""
    cols = list(estimate_df.columns)
    today = datetime.now().strftime("%Y. %m. %d.")
    hr = _title_block(
        ws, len(cols), "공 사 세 부 내 역 서 (을지)",
        [("공 사 명", project_name), ("작성일자", today), ("작 성 자", writer_name)]
    )
    _write_header(ws, hr, cols)

    num_cols = []
    r = hr + 1
    for i, (_, rec) in enumerate(estimate_df.iterrows()):
        for j, cname in enumerate(cols, start=1):
            val = _to_number(rec[cname])
            if isinstance(val, float) and pd.isna(val):
                val = None
            c = ws.cell(row=r, column=j, value=val)
            c.border = BOX
            c.font = Font(name=FONT_NAME, size=10, color="333333")
            fmt = _fmt_for(cname)
            if fmt and isinstance(val, (int, float)):
                c.number_format = fmt
                c.alignment = Alignment(horizontal="right", vertical="center")
                if any(k in str(cname) for k in ("금액", "합계", "단가")):
                    num_cols.append(j)
            elif j == 1 or str(cname) in ("구분", "단위", "비고"):
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if i % 2 == 1:
                c.fill = PatternFill("solid", start_color=C_ZEBRA_BG)
        ws.row_dimensions[r].height = 19
        r += 1

    # 합계 행 : Excel 수식(SUM) 사용 → 데이터 수정 시 자동 재계산
    if len(estimate_df) > 0:
        first_data, last_data = hr + 1, r - 1
        tc = ws.cell(row=r, column=1, value="합    계")
        tc.font = Font(name=FONT_NAME, size=11, bold=True, color="1F3864")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        for j in range(1, len(cols) + 1):
            c = ws.cell(row=r, column=j)
            c.fill = PatternFill("solid", start_color=C_TOTAL_BG)
            c.border = Border(left=THIN, right=THIN, top=MEDIUM, bottom=MEDIUM)
            if j in set(num_cols) and any(k in str(cols[j - 1]) for k in ("금액", "합계")):
                L = get_column_letter(j)
                c.value = f"=SUM({L}{first_data}:{L}{last_data})"
                c.number_format = NUM_FMT
                c.font = Font(name=FONT_NAME, size=11, bold=True, color="1F3864")
                c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[r].height = 24
        ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(cols))}{last_data}"

    _autofit(ws, hr)
    _page_setup(ws, hr, landscape=len(cols) > 6)


def _build_rates_sheet(ws, rates):
    """적용 제비율 시트 (rates 전달 시에만 생성)"""
    labels = {
        "indirect_labor": "간접노무비율", "sanjae": "산재보험료율", "goyong": "고용보험료율",
        "health": "국민건강보험료율", "elderly": "노인장기요양보험료율", "pension": "국민연금보험료율",
        "retire": "퇴직공제부금비율", "safety": "산업안전보건관리비율", "env": "환경보전비율",
        "etc_exp": "기타경비율", "general_admin": "일반관리비율", "profit": "이윤율", "tax": "부가가치세율",
    }
    hr = _title_block(ws, 3, "적 용 제 비 율",
                      [("기준", "조달청 원가계산 제비율 적용기준"),
                       ("작성일자", datetime.now().strftime("%Y. %m. %d."))])
    _write_header(ws, hr, ["구분", "적용요율(%)", "비고"])

    r = hr + 1
    for k, v in rates.items():
        lc = ws.cell(row=r, column=1, value=labels.get(k, k))
        lc.font = Font(name=FONT_NAME, size=10, bold=True)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        vc = ws.cell(row=r, column=2, value=float(v))
        vc.number_format = '0.00"%"'
        vc.font = Font(name=FONT_NAME, size=10, color="0000FF")   # 입력값 = 청색
        vc.alignment = Alignment(horizontal="center", vertical="center")
        nc = ws.cell(row=r, column=3, value="")
        nc.font = Font(name=FONT_NAME, size=9, color=C_NOTE)
        for j in range(1, 4):
            ws.cell(row=r, column=j).border = BOX
        ws.row_dimensions[r].height = 18
        r += 1

    _autofit(ws, hr)
    _page_setup(ws, hr)


# ═══════════════════════════════════════════════════════════
# 4. 엑셀 바이너리 생성 (외부 호출 진입점)
# ═══════════════════════════════════════════════════════════
def generate_excel_bytes(estimate_df, summary_df, project_name="공사 견적",
                         writer_name="시설관리팀", rates=None, note=None):
    """
    세부내역서(을지) + 원가계산서(갑지) [+ 적용제비율] 시트를 포함한
    서식 적용 엑셀 바이너리 생성

    estimate_df  : DataFrame  세부내역서 원본
    summary_df   : DataFrame  calculate_cost_summary() 결과 DataFrame
    project_name : str        공사명 (제목 블록)
    writer_name  : str        작성자/부서
    rates        : dict|None  전달 시 '적용제비율' 시트 추가
    note         : str|None   갑지 하단 각주 문구 교체
    """
    if estimate_df is None:
        estimate_df = pd.DataFrame()
    if summary_df is None:
        summary_df = pd.DataFrame()

    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("원가계산서(갑지)")
    if not summary_df.empty:
        _build_summary_sheet(ws1, summary_df, project_name, writer_name, note)
    ws1.sheet_properties.tabColor = C_TITLE_BG

    ws2 = wb.create_sheet("세부내역서(을지)")
    if not estimate_df.empty:
        _build_estimate_sheet(ws2, estimate_df, project_name, writer_name)
    ws2.sheet_properties.tabColor = C_HEADER_BG

    if rates:
        ws3 = wb.create_sheet("적용제비율")
        _build_rates_sheet(ws3, rates)
        ws3.sheet_properties.tabColor = "A6A6A6"

    wb.properties.title = f"{project_name} 원가계산서"
    wb.properties.creator = writer_name

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()